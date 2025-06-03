import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from snowflake_connection import get_snowflake_connection


def get_column_types(conn, database, schema, table):
    query = f"""
        SELECT COLUMN_NAME, DATA_TYPE
        FROM {database}.INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table}'
    """
    cur = conn.cursor()
    cur.execute(query)
    rows = cur.fetchall()
    cur.close()

    categorical = [r[0] for r in rows if r[1].upper() in ('TEXT', 'VARCHAR', 'CHAR', 'STRING')]
    numeric = [r[0] for r in rows if r[1].upper() in ('NUMBER', 'FLOAT', 'INT', 'DECIMAL', 'NUMERIC', 'DOUBLE')]
    return categorical, numeric


def get_skew_data_with_query(conn, database, schema, table, columns, top_n=3, skew_threshold=0.8):
    cur = conn.cursor()
    summary = []

    for col in columns:
        query = f"""
            SELECT "{col}", COUNT(*) as cnt
            FROM "{database}"."{schema}"."{table}"
            GROUP BY "{col}"
            ORDER BY cnt DESC
            LIMIT {top_n}
        """.strip()

        cur.execute(query)
        rows = cur.fetchall()

        total_query = f'SELECT COUNT(*) FROM "{database}"."{schema}"."{table}"'
        cur.execute(total_query)
        total_rows = cur.fetchone()[0]

        top_value = rows[0][0] if rows else None
        top_count = rows[0][1] if rows else 0
        skew = "Yes" if total_rows > 0 and (top_count / total_rows) > skew_threshold else "No"

        summary.append({
            "Column": col,
            "Top_Value": top_value,
            "Top_Count": top_count,
            "Total_Rows": total_rows,
            "Dominance_%": f"{(top_count / total_rows) * 100:.2f}%" if total_rows > 0 else "0%",
            "Skew_Detected": skew,
            "Query_Used": query
        })

    cur.close()
    return pd.DataFrame(summary)


def get_outlier_data_with_query(conn, database, schema, table, columns):
    cur = conn.cursor()
    summary = []

    for col in columns:
        query = f'SELECT "{col}" FROM "{database}"."{schema}"."{table}" WHERE "{col}" IS NOT NULL'
        df = pd.read_sql(query, conn)

        if df.empty:
            continue

        Q1 = df[col].quantile(0.25)
        Q3 = df[col].quantile(0.75)
        IQR = Q3 - Q1
        lower = Q1 - 1.5 * IQR
        upper = Q3 + 1.5 * IQR
        outliers = df[(df[col] < lower) | (df[col] > upper)][col].tolist()

        summary.append({
            "Column": col,
            "Q1": Q1,
            "Q3": Q3,
            "IQR": IQR,
            "Lower_Bound": lower,
            "Upper_Bound": upper,
            "Outlier_Count": len(outliers),
            "Sample_Outliers": ", ".join(map(str, outliers[:3])) if outliers else "None",
            "Query_Used": query
        })

    cur.close()
    return pd.DataFrame(summary)


def write_to_excel(df, wb, sheet_name):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        if any(row):
            ws.append(row)


def run_skew_and_outlier_validation(config_path, input_excel):
    conn, database, schema = get_snowflake_connection(config_path)
    table_df = pd.read_excel(input_excel)

    for _, row in table_df.iterrows():
        table = row['Table']
        print(f"\n Validating: {database}.{schema}.{table}")
        cat_cols, num_cols = get_column_types(conn, database, schema, table)

        skew_df = get_skew_data_with_query(conn, database, schema, table, cat_cols)
        outlier_df = get_outlier_data_with_query(conn, database, schema, table, num_cols)

        output_file = f"{table}_skew_outlier_check.xlsx"
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        write_to_excel(skew_df, wb, "skew_check")
        write_to_excel(outlier_df, wb, "outlier_check")
        wb.save(output_file)
        print(f" Results saved: {output_file}")

    conn.close()
    print("\n🔒 Snowflake connection closed.")

