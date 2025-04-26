import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sql_parser import extract_joins_from_sql
from join_validator import validate_joins_from_list
from snowflake_connection import get_snowflake_connection

def read_sql_file(file_path):
    with open(file_path, 'r') as f:
        return f.read()

def write_summary_to_excel(parsed_joins, validation_summary_df, spot_check_samples, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    # Parsed Joins Summary
    parsed_df = pd.DataFrame([{
        'Left_Table': join['left_table'],
        'Right_Table': join['right_table'],
        'Left_Join_Columns': ', '.join(join['left_keys']),
        'Right_Join_Columns': ', '.join(join['right_keys']),
        'Join_Type': join['join_type'],
        'Validation_Status': join.get('Validation_Status', 'Pending')
    } for join in parsed_joins])

    ws1 = wb.create_sheet("Parsed_Joins_Summary")
    for r in dataframe_to_rows(parsed_df, index=False, header=True):
        ws1.append(r)

    # Join Validation Summary
    ws2 = wb.create_sheet("Join_Validation_Summary")
    for r in dataframe_to_rows(validation_summary_df, index=False, header=True):
        ws2.append(r)

    # Spot Check Samples
    for idx, (join_info, spot_check_df) in enumerate(spot_check_samples, start=1):
        sheet_name = f"Spot_Check_{idx}"
        ws = wb.create_sheet(sheet_name)

        # Metadata
        ws.append(["Left Table", join_info['left_table']])
        ws.append(["Right Table", join_info['right_table']])
        ws.append(["Join Keys", ', '.join(join_info['left_keys'])])
        ws.append(["Validation Timestamp", pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([])

        # Spot check data
        for r in dataframe_to_rows(spot_check_df, index=False, header=True):
            ws.append(r)

    # Save file
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Excel report saved to: {output_path}")

if __name__ == "__main__":
    # Settings
    sql_input_path = "input/your_query.sql"
    output_excel_path = "output/Join_Validation_Report.xlsx"
    config_path = "config.json"

    # Step 1: Connect to Snowflake
    conn, *_ = get_snowflake_connection(config_path)

    try:
        # Step 2: Parse SQL to extract joins
        sql_text = read_sql_file(sql_input_path)
        parsed_joins = extract_joins_from_sql(sql_text)

        # Step 3: Validate extracted joins
        validation_results = []
        spot_check_samples = []

        for idx, join in enumerate(parsed_joins, start=1):
            if join['join_type'] == 'INNER JOIN':
                validation_df = validate_joins_from_list([join], conn)
                validation_result = validation_df.to_dict(orient='records')[0]
                validation_results.append(validation_result)

                # Spot Check Sample
                if validation_result['Validation_Status'] == 'Validated':
                    # Re-run spot check to capture sample data
                    try:
                        from join_validator import run_spot_check_validation
                        spot_check_df = run_spot_check_validation(conn, join['left_table'], join['right_table'], join['left_keys'], join['right_keys'])
                        spot_check_samples.append((join, spot_check_df))
                    except Exception as e:
                        print(f"⚠️ Failed to capture spot check for join {idx}: {e}")
            else:
                join['Validation_Status'] = 'Skipped - Non-Inner Join'
                validation_results.append(join)

        validation_summary_df = pd.DataFrame(validation_results)

        # Step 4: Write results to Excel
        write_summary_to_excel(parsed_joins, validation_summary_df, spot_check_samples, output_excel_path)

    finally:
        conn.close()
        print("🔒 Snowflake connection closed.")
