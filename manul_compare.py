import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from snowflake_connectivity import get_snowflake_connection

# Function to fetch DESCRIBE TABLE results from Snowflake
def fetch_table_data(conn, schema_name, table_name, columns):
    query = f"SELECT {', '.join(columns)} FROM {schema_name}.{table_name}"
    cur = conn.cursor()
    cur.execute(query)
    data = cur.fetchall()
    df = pd.DataFrame(data, columns=columns)
    cur.close()
    return df

# Function to compare the results of two tables based on column mapping
def compare_table_data(df1, df2, column_mapping):
    mismatches = []
    
    # Compare rows based on mapped columns
    for col1, col2 in column_mapping.items():
        df_mismatch = df1[df1[col1] != df2[col2]]
        if not df_mismatch.empty:
            mismatches.append((col1, col2, df_mismatch))
    
    return mismatches

# Function to read column mapping from Excel file
def read_column_mapping(excel_file):
    df = pd.read_excel(excel_file)
    
    # Create a dictionary of source and target columns
    column_mapping = dict(zip(df['Src_clmn name'], df['Trg_ Clmsname']))
    
    # Extract source and target table names
    src_table = df['Src_table'].iloc[0]
    trg_table = df['Trg_table'].iloc[0]
    
    return column_mapping, src_table, trg_table

# Function to highlight mismatches in Excel
def highlight_mismatches(writer, sheet_name, mismatches):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for mismatch in mismatches:
        worksheet[f"A{mismatch[0] + 2}"].fill = fill
        worksheet[f"B{mismatch[0] + 2}"].fill = fill

# Main function to run the comparison and write results to Excel and console
def main():
    # Get connection from JSON config
    conn, _, schema_name, _, _, _, _ = get_snowflake_connection('config.json')

    # Read column mapping from Excel file
    excel_file = 'mapping_data.xlsx'  # Path to your Excel file
    column_mapping, src_table, trg_table = read_column_mapping(excel_file)

    # Fetch data from both tables
    df_src_table = fetch_table_data(conn, schema_name, src_table, list(column_mapping.keys()))
    df_trg_table = fetch_table_data(conn, schema_name, trg_table, list(column_mapping.values()))

    # Compare tables based on column mapping
    mismatches = compare_table_data(df_src_table, df_trg_table, column_mapping)

    # Output file for the results
    output_file = 'table_comparison_results.xlsx'

    # Write the results to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_src_table.to_excel(writer, sheet_name=f'{src_table}_Data', index=False)
        df_trg_table.to_excel(writer, sheet_name=f'{trg_table}_Data', index=False)

        # Highlight mismatches in both sheets
        highlight_mismatches(writer, f'{src_table}_Data', mismatches)
        highlight_mismatches(writer, f'{trg_table}_Data', mismatches)

    # Print mismatches in console
    if mismatches:
        print("\nMismatches found:")
        for mismatch in mismatches:
            col1, col2, df_mismatch = mismatch
            print(f"\nMismatch between {col1} and {col2}:")
            print(df_mismatch)
    else:
        print("\nNo mismatches found.")

    # Close the Snowflake connection
    conn.close()

if __name__ == "__main__":
    main()
